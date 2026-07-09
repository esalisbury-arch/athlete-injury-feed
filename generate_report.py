#!/usr/bin/env python3
"""Builds injury_report.xlsx from injuries.json."""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from injury_data import BASE_DIR, load_valid_records

REPORT_FILE = BASE_DIR / "injury_report.xlsx"

COLUMNS = [
    ("Athlete", "athlete", 20),
    ("Organization", "org", 30),
    ("Sport", "sport", 14),
    ("Injury", "injury_context", 40),
    ("Expected duration", "duration", 30),
    ("Financial impact", "financial_impact", 30),
    ("Team/game impact", "team_impact", 40),
    ("Date reported", "date_reported", 14),
    ("Source", "source_title", 30),
    ("Source URL", "source_url", 40),
]


def main() -> None:
    records, total = load_valid_records()
    records.sort(key=lambda r: r["date_reported"], reverse=True)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Injuries"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[cell.column_letter].width = width

    body_font = Font(name="Arial")
    for row_idx, record in enumerate(records, start=2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=record[key])
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row=row_idx, column=len(COLUMNS)).hyperlink = record["source_url"]

    sheet.freeze_panes = "A2"
    last_row = max(len(records) + 1, 2)
    last_col_letter = sheet.cell(row=1, column=len(COLUMNS)).column_letter
    table = Table(displayName="Injuries", ref=f"A1:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)

    tmp_file = REPORT_FILE.with_suffix(".xlsx.tmp")
    wb.save(tmp_file)
    os.replace(tmp_file, REPORT_FILE)
    print(f"Wrote {len(records)} of {total} rows to {REPORT_FILE}")


if __name__ == "__main__":
    main()
